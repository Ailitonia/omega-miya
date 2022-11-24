from pathlib import Path
from json import dumps
from openpyxl import load_workbook


def convert_xlsx_to_py(xlsx_file: Path) -> bool:
    """老黄历xlsx文件转为py

    Args:
        xlsx_file (Path): xlsx文件路径

    Returns:
        bool: False: 未出错
    """
    if not xlsx_file.exists():
        return True
    wbook = load_workbook(xlsx_file)
    wsheet = wbook['sheet']
    things = []
    for row in wsheet.iter_rows(min_row=2):
        things.append({
            'name': row[0].value,
            'good': row[1].value,
            'bad': row[2].value
        })
    with open('do_something.py', 'w', encoding='utf-8') as py_file:
        py_file.write('do_something = ' +
                      dumps(things, ensure_ascii=False, indent=4) + '\n')


if __name__ == "__main__":
    xlsx = Path.cwd().joinpath('tool_make_do_something.xlsx')
    if convert_xlsx_to_py(xlsx):
        print('An error has occurred.')
