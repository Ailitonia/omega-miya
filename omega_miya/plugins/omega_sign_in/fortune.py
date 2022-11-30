import random
import hashlib
import ujson as json
from datetime import datetime
from typing import Literal
from pydantic import BaseModel, parse_obj_as
from openpyxl import load_workbook, Workbook

from omega_miya.service.omega_api import register_get_route, BaseApiModel, BaseApiReturn
from omega_miya.utils.process_utils import run_sync

from .config import sign_in_config, sign_local_resource_config


class Fortune(BaseModel):
    """求签结果"""
    star: str
    text: Literal['大凶', '末凶', '半凶', '小凶', '凶', '末小吉', '末吉', '半吉', '吉', '小吉', '中吉', '大吉']
    good_do_st: str
    good_do_nd: str
    bad_do_st: str
    bad_do_nd: str


class FortuneEvent(BaseModel):
    """老黄历宜和不宜内容"""
    name: str
    good: str
    bad: str

    def __eq__(self, other) -> bool:
        if isinstance(other, FortuneEvent):
            return self.name == other.name and self.good == other.good and self.bad == other.bad
        else:
            return False

    def __hash__(self) -> int:
        return hash(self.name + self.good + self.bad)


__fortune_event: list[FortuneEvent] = []


def _upgrade_fortune_event(*, enforce_origin: bool = False) -> None:
    """更新求签事件缓存

    :param enforce_origin: 强制使用原始资源解析事件
    """
    if sign_local_resource_config.tmp_fortune_event.is_file and not enforce_origin:
        event_file = sign_local_resource_config.tmp_fortune_event
    else:
        event_file = sign_local_resource_config.default_fortune_event

    with event_file.open('r', encoding='utf8') as f:
        fortune_event = json.loads(f.read())

    global __fortune_event
    __fortune_event = parse_obj_as(list[FortuneEvent], fortune_event)


def get_fortune_event() -> list[FortuneEvent]:
    """获取所有求签事件"""
    if not __fortune_event:
        _upgrade_fortune_event()

    return __fortune_event


def random_fortune_event(num: int = 4) -> list[FortuneEvent]:
    """随机获取求签事件"""
    return random.sample(get_fortune_event(), k=num)


def _load_fortune_event_import_file() -> None:
    """从 xlsx 导入文件中读取并转换格式"""
    if not sign_local_resource_config.fortune_event_import_file.is_file:
        raise FileNotFoundError('Fortune event import file not found')

    import_workbook = load_workbook(sign_local_resource_config.fortune_event_import_file.path)
    import_sheet = import_workbook['fortune_event']
    fortune_event = []
    for row in import_sheet.iter_rows(min_row=2):
        fortune_event.append(
            FortuneEvent.parse_obj({
                'name': row[0].value,
                'good': row[1].value,
                'bad': row[2].value
            })
        )

    fortune_event = set(fortune_event)
    _upgrade_fortune_event(enforce_origin=True)
    fortune_event.update(__fortune_event)
    fortune_event = sorted([x.dict() for x in fortune_event], key=lambda x: x['name'])

    with sign_local_resource_config.tmp_fortune_event.open('w', encoding='utf8') as f:
        json.dump(fortune_event, f, ensure_ascii=False)


def _create_fortune_event_import_file_template() -> None:
    """生成空白的 xlsx 导入文件模板"""
    template_workbook = Workbook()
    template_sheet = template_workbook.active
    template_sheet.title = 'fortune_event'

    template_sheet.cell(column=1, row=1, value='事项')
    template_sheet.cell(column=2, row=1, value='吉')
    template_sheet.cell(column=3, row=1, value='凶')

    with sign_local_resource_config.fortune_event_import_file.open('wb') as f:
        template_workbook.save(f)


@register_get_route(path='/fortune/import-fortune-event', enabled=sign_in_config.signin_enable_fortune_import_api)
async def _handle_fortune_event_import(reset: bool = False) -> BaseApiReturn:
    """api 导入求签事件

    :param reset: 重置求签事件为默认值(调试用, 仅影响事件缓存)
    """
    class _ReturnBody(BaseApiModel):
        events: list[FortuneEvent]

    class _Return(BaseApiReturn):
        body: _ReturnBody

    # 没有导入文件则生成导入文件模板
    if not sign_local_resource_config.fortune_event_import_file.is_file:
        try:
            await run_sync(_create_fortune_event_import_file_template)()
            return _Return(
                error=False,
                body=_ReturnBody(events=[]),
                message=f'未找到导入文件, '
                        f'已生成导入模板: {sign_local_resource_config.fortune_event_import_file.resolve_path}, '
                        f'请在模板上直接修改或覆盖后重新导入'
            )
        except Exception as e:
            return _Return(
                error=True,
                body=_ReturnBody(events=[]),
                message=f'创建导入模板失败: {sign_local_resource_config.fortune_event_import_file.resolve_path}, '
                        f'请检查文件路径是否被占用',
                exception=str(e)
            )

    # 有导入文件则读取并解析导入文件
    try:
        await run_sync(_load_fortune_event_import_file)()
        await run_sync(_upgrade_fortune_event)(enforce_origin=reset)
        return _Return(
            error=False,
            body=_ReturnBody(events=get_fortune_event()),
            message=f'求签事件导入并更新成功'
        )
    except Exception as e:
        return _Return(
            error=True,
            body=_ReturnBody(events=[]),
            message=f'求签事件导入或更新失败',
            exception=str(e)
        )


def get_fortune(user_id: int, *, date: datetime | None = None) -> Fortune:
    """根据 qq 号和当天日期生成老黄历"""
    if date is None:
        date_str = str(datetime.today().date())
    else:
        date_str = str(date.date())
    # 用qq、日期生成随机种子
    random_seed_str = str([user_id, date_str])
    md5 = hashlib.md5()
    md5.update(random_seed_str.encode('utf-8'))
    random_seed = md5.hexdigest()
    random.seed(random_seed)
    # 今日运势
    # 生成运势随机数
    fortune_result = random.randint(1, 108)
    # 大吉・中吉・小吉・吉・半吉・末吉・末小吉・凶・小凶・半凶・末凶・大凶
    if fortune_result < 4:
        fortune_star = '☆' * 11
        fortune_text = '大凶'
    elif fortune_result < 9:
        fortune_star = '★' * 1 + '☆' * 10
        fortune_text = '末凶'
    elif fortune_result < 16:
        fortune_star = '★' * 2 + '☆' * 9
        fortune_text = '半凶'
    elif fortune_result < 25:
        fortune_star = '★' * 3 + '☆' * 8
        fortune_text = '小凶'
    elif fortune_result < 36:
        fortune_star = '★' * 4 + '☆' * 7
        fortune_text = '凶'
    elif fortune_result < 48:
        fortune_star = '★' * 5 + '☆' * 6
        fortune_text = '末小吉'
    elif fortune_result < 60:
        fortune_star = '★' * 6 + '☆' * 5
        fortune_text = '末吉'
    elif fortune_result < 72:
        fortune_star = '★' * 7 + '☆' * 4
        fortune_text = '半吉'
    elif fortune_result < 84:
        fortune_star = '★' * 8 + '☆' * 3
        fortune_text = '吉'
    elif fortune_result < 96:
        fortune_star = '★' * 9 + '☆' * 2
        fortune_text = '小吉'
    elif fortune_result < 102:
        fortune_star = '★' * 10 + '☆' * 1
        fortune_text = '中吉'
    else:
        fortune_star = '★' * 11
        fortune_text = '大吉'

    # 宜做和不宜做
    do_and_not = random_fortune_event(num=4)

    result = {
        'star': fortune_star,
        'text': fortune_text,
        'good_do_st': f"{do_and_not[0].name} —— {do_and_not[0].good}",
        'good_do_nd': f"{do_and_not[2].name} —— {do_and_not[2].good}",
        'bad_do_st': f"{do_and_not[1].name} —— {do_and_not[1].bad}",
        'bad_do_nd': f"{do_and_not[3].name} —— {do_and_not[3].bad}"
    }

    # 重置随机种子
    random.seed()

    return Fortune.parse_obj(result)


__all__ = [
    'get_fortune'
]
